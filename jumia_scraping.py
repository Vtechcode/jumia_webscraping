import asyncio
import requests
from bs4 import BeautifulSoup
from urllib.parse import quote_plus
import time
import aiohttp
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import os

# checking if the excel workbook exists
if not os.path.exists('jumia_webscraping.xlsx'):
    # Creating a new workbook if it doesn't exist
    wb = Workbook()
    ws = wb.active
    ws.title = 'Prices'

    headings = ['Search Name', 'Product', 'Prce',]

    ws.append(headings)

else:
    # Loading an existing workbook
    wb = load_workbook('jumia_webscraping.xlsx')
    ws = wb.active

# coroutine to scrape the prices from Jumia website
async def scraping(url, searched_item):
     async with aiohttp.ClientSession() as session:
        async with session.get(url) as response:
            reply = await response.text()
            soup = BeautifulSoup(reply, 'lxml')
            info_div = soup.find('div', {'class': 'info'})
            name = info_div.find('h3', {'class': 'name'})
            price = info_div.find('div', {'class': 'prc'})
            return [searched_item, name.text.strip(), price.text.strip()]

        
start = time.time()

# coroutine to create tasks for scraping items prices
async def get_items_and_url():
    jumia_url = 'https://www.jumia.co.ke/catalog/?q={}'

    items = []

    # Gathering user input for items to search
    for i in range(5):
        item = input("Item: ")
        items.append(item)

    tasks = []
    for item in items:
        url = jumia_url.format(quote_plus(item))
        task = asyncio.create_task(scraping(url, item))
        tasks.append(task)
   
    # awaiting for all the tasks to be complete
    content = await asyncio.gather(*tasks)

    # Inputting the data into the Excel spreadsheet
    for data in content:
        specific_to_be_put_into_excel = [data[0], data[1], data[2]]
        ws.append(specific_to_be_put_into_excel)

    wb.save('jumia_webscraping.xlsx')

    print(content)


    
total_time = time.time() - start

if __name__ == "__main__":
    loop = asyncio.get_event_loop()
    loop.run_until_complete(get_items_and_url())

print(total_time)